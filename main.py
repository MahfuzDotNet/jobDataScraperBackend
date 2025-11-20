from flask import Flask, request, session, jsonify, send_file, send_from_directory
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
import logging
import cloudscraper



app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')

# Configure CORS for cross-origin requests - Allow all origins in development
# CORS(app, 
#      supports_credentials=True,
#      resources={r"/api/*": {"origins": "*"}},
#      allow_headers=['Content-Type', 'Authorization'],
#      expose_headers=['Content-Type'],
#      methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])


CORS(app,
     supports_credentials=True,
     resources={r"/api/*": {"origins": ["http://192.168.0.102:5000"]}},
     allow_headers=["Content-Type", "Authorization"],
     expose_headers=["Content-Type"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])




# Configure session cookie settings for browser compatibility
app.config['SESSION_COOKIE_SECURE'] = False  # False for development (HTTP), True for production (HTTPS)
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Keep HttpOnly for security
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Lax for better cross-browser compatibility

# app.config['SESSION_COOKIE_SECURE'] = False  # keep False for localhost
# app.config['SESSION_COOKIE_HTTPONLY'] = True
# app.config['SESSION_COOKIE_SAMESITE'] = 'None'  # must be None for cross-origin cookies



# Sample job data for demonstration
sample_jobs = []


@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    valid_credentials = {  
        'demo': 'demo123'
    }

    if username and password:
        if username in valid_credentials and valid_credentials[username] == password:
            session['username'] = username
            return jsonify({'success': True, 'username': username})
        else:
            return jsonify({'success': False, 'error': 'Invalid username or password'}), 401
    else:
        return jsonify({'success': False, 'error': 'Please enter both username and password'}), 400


@app.route('/api/logout', methods=['POST'])
def logout():
    session.pop('username', None)
    return jsonify({'success': True})

@app.route('/api/me')
def get_user():
    if 'username' not in session:
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, 'username': session['username']})



@app.route('/api/scrape', methods=['POST'])
def scrape_jobs():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    url_to_scrape = data.get('url_to_scrape')
    logging.debug("url_to_scrape:", url_to_scrape)
    max_pages = int(data.get('max_pages', 1))
    category = data.get('category')

    session['scraping_url'] = url_to_scrape
    session['scraping_pages'] = max_pages
    session['scraping_category'] = category

    global sample_jobs
    sample_jobs = []

    scraped_data = scrape_website_content(url_to_scrape, max_pages)
    sample_jobs.extend(scraped_data)

    return jsonify({'success': True, 'jobs': sample_jobs, 'count': len(sample_jobs)})



import re
from urllib.parse import urljoin
from datetime import datetime, timedelta

def scrape_website_content(url, max_pages=1):
    """
    Entry function called by your /api/scrape endpoint.
    Routes to the appropriate site-specific scraper based on domain.
    """
    domain = ''
    try:
        domain = re.sub(r'^https?://(www\.)?', '', url).split('/')[0].lower()
    except Exception:
        domain = ''

    if 'kosovajob' in domain:
        return scrape_kosovajob(url, max_pages=max_pages)
    elif 'ofertapune' in domain:
        return scrape_ofertapune(url, max_pages=max_pages)
    elif 'telegrafi' in domain:
        return scrape_telegrafi(url, max_pages=max_pages)
    elif 'portalpune' in domain:
        return scrape_portalpune(url, max_pages=max_pages)
    else:
        return generic_scrape_site(url, max_pages=max_pages)



from playwright.sync_api import sync_playwright
from playwright_stealth import stealth_sync
from bs4 import BeautifulSoup
import time, random, re
from datetime import datetime


# ----------------- DETAIL PAGE PARSER ------------------

def parse_kosovajob_detail(job_url, page, base_url):
    """
    Robust KosovaJob detail parser.
    Tries many selectors / fallbacks to reliably return company & location.
    """
    try:
        # navigate and give JS a short moment to render key blocks
        page.goto(job_url, wait_until="networkidle", timeout=60000)
        # optional small wait to let images/JS populate
        try:
            page.wait_for_selector("body", timeout=3000)
        except:
            pass
    except Exception as e:
        print("Detail blocked:", job_url, e)
        return None

    soup = BeautifulSoup(page.content(), "html.parser")

    def txt(sel):
        return sel.get_text(strip=True, separator=" ") if sel else ""

    # ---------- TITLE ----------
    title_selectors = [
        "h1[itemprop='title']",
        "h1.jobTitle",
        "h1",
        ".jobListTitle",
        ".jobTitle",
        "div.jobListCntsInner .jobListTitle"
    ]
    title = ""
    for sel in title_selectors:
        el = soup.select_one(sel)
        if el and el.get_text(strip=True):
            title = el.get_text(strip=True)
            break
    title = title or ""

    # ---------- COMPANY (robust) ----------
    company = ""

    # 1) Common dedicated selectors (detail page structures)
    company_selectors = [
        ".companySinglePage img[alt]",
        ".companySinglePage",
        ".companySinglePage .companyName",
        ".jobCompanyBox img[alt]",
        ".company, .company-name, .employer, .companySinglePage .company",
        "div.companySinglePage"
    ]
    for sel in company_selectors:
        el = soup.select_one(sel)
        if not el:
            continue
        # if element is an <img> with alt
        if el.name == "img" and el.get("alt"):
            company = el.get("alt").strip()
            break
        # if it's a container, try to find an <img alt> inside first
        img = el.select_one("img[alt]")
        if img and img.get("alt"):
            company = img.get("alt").strip()
            break
        # otherwise try to get textual content from the container
        text = txt(el)
        if text and len(text) > 1 and "http" not in text.lower():
            company = text.strip()
            break

    # 2) Scan all <img> and prefer ones in parents with 'company'/'logo' keywords
    if not company:
        imgs = soup.find_all("img")
        best = None
        for img in imgs:
            alt = (img.get("alt") or "").strip()
            parent_classes = " ".join(img.parent.get("class") or [])
            parent_text = txt(img.parent)
            score = 0
            if alt:
                score += 10
            if re.search(r"(company|logo|employ|firma|employer|company-name)", parent_classes, re.I):
                score += 8
            if re.search(r"(company|logo|employer|firma)", parent_text, re.I):
                score += 5
            if score > 0:
                # prefer higher score
                if not best or score > best[0]:
                    best = (score, img, alt)
        if best:
            alt = best[2]
            if alt:
                company = alt.strip()

    # 3) Try JSON-LD (structured data) for hiringOrganization / author
    if not company:
        for script in soup.select("script[type='application/ld+json']"):
            try:
                import json
                data = json.loads(script.string or "{}")
                # data might be dict or list
                if isinstance(data, list):
                    candidates = data
                else:
                    candidates = [data]
                for obj in candidates:
                    # check hiringOrganization
                    ho = obj.get("hiringOrganization") if isinstance(obj, dict) else None
                    if isinstance(ho, dict):
                        name = ho.get("name")
                        if name:
                            company = name.strip()
                            break
                    # check author
                    author = obj.get("author")
                    if isinstance(author, dict):
                        name = author.get("name")
                        if name:
                            company = name.strip()
                            break
                if company:
                    break
            except Exception:
                continue

    # 4) Final fallback: parse company slug from the job_url path (works often)
    if not company:
        try:
            from urllib.parse import urlparse
            path = urlparse(job_url).path.strip("/")
            parts = [p for p in path.split("/") if p]
            # common pattern: /company-slug/job-slug   -> take first part
            if parts:
                candidate = parts[0]
                # skip if candidate looks like 'punet' or numeric
                if candidate and not candidate.isnumeric() and candidate.lower() not in ("punet", "jobs"):
                    company = candidate.replace("-", " ").replace("_", " ").title()
        except Exception:
            pass

    # ensure company not empty
    if not company:
        company = "N/A"

  
    # ---------- LOCATION (robust) ----------
    location = ""

    # 1) Common detail page selectors
    loc_selectors = [
        ".jobInfoBox .jobInfoText",
        ".jobInfoText",
        ".jobLocation",
        ".location",
        ".jobListCity",  # listing page support
    ]

    for sel in loc_selectors:
        el = soup.select_one(sel)
        if el:
            txtv = el.get_text(strip=True)
            if txtv:
                location = txtv
                break

    # 2) Listing-page fallback: find `.jobListCity` anywhere
    if not location:
        city = soup.find(class_="jobListCity")
        if city:
            txtv = city.get_text(strip=True)
            if txtv:
                location = txtv

    # 3) Fallback: search inside job containers (.jobListCnts or .jobListCntsInner)
    if not location:
        containers = soup.select(".jobListCnts, .jobListCntsInner")
        for c in containers:
            city = c.find(class_="jobListCity")
            if city:
                txtv = city.get_text(strip=True)
                if txtv:
                    location = txtv
                    break

    # 4) Try meta tags
    if not location:
        meta_loc = soup.select_one(
            "meta[name='location'], meta[property='jobLocation'], meta[property='og:locality']"
        )
        if meta_loc and meta_loc.get("content"):
            location = meta_loc.get("content").strip()

    # 5) Final fallback
    if not location:
        location = "Kosovo"





    # ---------- other fields ----------
    posted_on = txt(soup.find("time")) or datetime.utcnow().strftime("%Y-%m-%d")

    desc = txt(soup.select_one(".job-description, .description, .content, .jobDesc"))
    if not desc:
        # try a large container commonly used for job text
        big = soup.select_one("#jobDetails, .jobDetails, .jobSingle, .job-page")
        desc = txt(big) if big else soup.get_text(" ", strip=True)[:2000]

    page_text = soup.get_text(" ", strip=True)

    email = re.search(r'[\w\.-]+@[\w\.-]+\.\w{2,}', page_text)
    phone = re.search(r'\+?\d[\d\s-]{5,}', page_text)
    salary = re.search(r'\d{1,3}(?:[.,]\d{3})?\s*(EUR|€)', page_text, re.I)
    jtype = re.search(r'(Full[-\s]*time|Part[-\s]*time|Intern|Contract)', page_text, re.I)
    job_id = re.search(r'(\d{4,})', job_url)

    return {
        "company": company,
        "position": title,
        "phone_number": phone.group(0) if phone else "",
        "email": email.group(0) if email else "",
        "category": txt(soup.select_one(".breadcrumb")) or "General",
        "posting_date": posted_on,
        "scraping_date": datetime.utcnow().strftime("%Y-%m-%d"),
        "location": location,
        "job_link": job_url,
        "description": desc,
        "source": "KosovaJob",
        "website": base_url,
        "salary": salary.group(0) if salary else "",
        "job_type": jtype.group(0).title() if jtype else "",
        "jobpost_ad_id": job_id.group(0) if job_id else ""
    }




# ------------------ FULL SCRAPER ------------------
def scrape_kosovajob(base_url="https://www.kosovajob.com", max_pages=None, max_scrolls=20, delay=2.5):

    if max_pages is not None:
        max_scrolls = max_pages

    scraped_jobs = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        page = browser.new_page()
        stealth_sync(page)

        print("\n=== Loading job list (/punet) ===")
        # page.goto(base_url + "/punet", timeout=60000)

        page.goto(base_url, timeout=60000)

        # Wait for new CSS selector
        try:
            page.wait_for_selector("div.jobListCnts", timeout=20000)
            print("✓ Job list rendered")
        except:
            print("✗ List did NOT render! Blocked or wrong selector.")
            browser.close()
            return scraped_jobs

        # SCROLL TO LOAD MORE JOBS
        last_height = None
        for _ in range(max_scrolls):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(delay)
            new_height = page.evaluate("document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # Parse HTML
        soup = BeautifulSoup(page.content(), "html.parser")
        cards = soup.select("div.jobListCnts")      # updated selector
        print("Total cards found:", len(cards))

        links = set()
        for card in cards:
            a = card.select_one("a")
            if a and a.get("href"):
                links.add(a["href"])

        print("Total job links extracted:", len(links))

        # Load detail pages
        for job_url in links:
            time.sleep(0.5 + random.random())
            job = parse_kosovajob_detail(job_url, page, base_url)
            if job:
                scraped_jobs.append(job)
                print("✓", job["position"])
            else:
                print("✗ Failed:", job_url)

        browser.close()

    return scraped_jobs







def generic_scrape_site(url, max_pages=1):
    """Fallback generic scraper for non-kosovajob URLs"""
    print(f"[INFO] Using generic scraper for {url}")
    return generate_sample_jobs()


from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import random
import os

# Set up user-agent rotation
user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.3",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/18.17763",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
]


import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import time
import random

def scrape_ofertapune(base_url="https://ofertapune.net", max_pages=None):
    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--start-maximized")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3")

    driver = uc.Chrome(options=options)

    print("=== Opening OfertaPune Browser ===")

    driver.get(base_url)
    time.sleep(5)

    # Handle Captcha
    try:
        captcha = driver.find_element("xpath", "//iframe[contains(@src, 'captcha')]")
        input("Please solve the captcha and press Enter...")
    except:
        pass

    # SCROLL HUMAN-LIKE
    for _ in range(8):
        driver.execute_script("window.scrollBy(0, document.body.scrollHeight);")
        time.sleep(1 + random.random())

    soup = BeautifulSoup(driver.page_source, "html.parser")

    # REAL OfertaPune posts
    cards = soup.select("article, .post, .col-md-4")

    if not cards:
        print("✗ Still blocked — but only if the site shows a captcha page.")
        print("Solve the captcha manually → keep browser open → run again.")
        driver.quit()
        return []

    print(f"✓ Found {len(cards)} posts on homepage")

    # Extract links
    links = set()
    for c in cards:
        a = c.select_one("a[href*='/jobs/']")
        if a:
            href = a["href"]
            if href.startswith("/"):
                href = base_url + href
            links.add(href)

    print("✓ Detail links:", len(links))

    scraped_jobs = []

    # LOAD EACH DETAIL PAGE
    for url in links:
        try:
            driver.get(url)
            time.sleep(2)

            html = driver.page_source
            soup = BeautifulSoup(html, "html.parser")

            # Use your SAME parser
            job = parse_kosovajob_detail(url, soup, base_url)

            if job:
                scraped_jobs.append(job)
                print("✓", job["position"])
        except Exception as e:
            print("Error:", e)

    driver.quit()
    return scraped_jobs




# file: telegrafi_playwright_scraper.py
"""
Telegrafi scraper using Playwright and a dedicated parse_telegrafi_detail(...) function.
Requirements:
    pip install playwright beautifulsoup4
    playwright install
Run:
    python telegrafi_playwright_scraper.py
"""
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urljoin, urlparse
import time
import random
import re
import os

# -----------------------------------------
# Minimal UA pool and stealth helper
# -----------------------------------------
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
]


def stealth_sync(page):
    """Small stealthing measures; reduces trivial automation fingerprints."""
    ua = random.choice(USER_AGENTS)
    try:
        page.context.set_user_agent(ua)
    except Exception:
        pass
    page.set_extra_http_headers({"accept-language": "en-US,en;q=0.9"})
    page.context.set_default_navigation_timeout(60000)
    page.add_init_script(
        """
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
        Object.defineProperty(navigator, 'languages', {get: () => ['en-US','en']});
        window.chrome = { runtime: {} };
        const orig = navigator.permissions.query;
        navigator.permissions.query = (p) => p.name === 'notifications' ? Promise.resolve({ state: Notification.permission }) : orig(p);
        """
    )


# -----------------------------------------
# Robust detail parser for Telegrafi
# -----------------------------------------
def parse_telegrafi_detail(job_url: str, page, base_url: str):
    """
    Parse a Telegrafi job detail page using the given Playwright `page`.
    Returns a dict with standardized fields or None on failure.
    """
    try:
        page.goto(job_url, wait_until="domcontentloaded", timeout=60000)
        # give JS a short moment to populate interactive content
        try:
            page.wait_for_selector("body", timeout=3000)
        except PWTimeout:
            pass
        time.sleep(0.5 + random.random() * 0.7)
    except Exception as e:
        print("Detail navigate error:", job_url, e)
        return None

    soup = BeautifulSoup(page.content(), "html.parser")

    def txt(el):
        return el.get_text(strip=True, separator=" ") if el else ""

    # ---------- TITLE ----------
    title_selectors = [
        "h1.entry-title",
        "h1.post-title",
        "h1",
        "h2.text-lg",
        ".post-title",
        ".entry-title"
    ]
    title = ""
    for sel in title_selectors:
        el = soup.select_one(sel)
        if el and el.get_text(strip=True):
            title = el.get_text(strip=True)
            break
    title = title or ""

    # ---------- COMPANY (robust) ----------
    company = ""
    # 1) direct badge text (Telegrafi snippet uses span.font-medium.text-gray-600)
    el = soup.select_one("span.font-medium.text-gray-600")
    if el and el.get_text(strip=True):
        company = el.get_text(strip=True)

    # 2) dedicated selectors / containers
    if not company:
        company_selectors = [
            ".company, .post-company, .job-company, .company-name",
            ".single-company, header .company", ".org-name"
        ]
        for sel in company_selectors:
            el = soup.select_one(sel)
            if el:
                # prefer image alt inside container
                img = el.select_one("img[alt]")
                if img and img.get("alt"):
                    company = img.get("alt").strip()
                    break
                text = txt(el)
                if text and len(text) > 1 and "http" not in text.lower():
                    company = text.strip()
                    break

    # 3) scan images heuristics (logo alt)
    if not company:
        imgs = soup.find_all("img")
        best = None
        for img in imgs:
            alt = (img.get("alt") or "").strip()
            parent_classes = " ".join(img.parent.get("class") or [])
            parent_text = txt(img.parent)
            score = 0
            if alt:
                score += 8
            if re.search(r"(company|logo|employ|firma|organization)", parent_classes, re.I):
                score += 6
            if re.search(r"(company|logo|employer|firma|org)", parent_text, re.I):
                score += 3
            if score > 0 and (not best or score > best[0]):
                best = (score, img, alt)
        if best and best[2]:
            company = best[2].strip()

    # 4) JSON-LD (structured data)
    if not company:
        for script in soup.select("script[type='application/ld+json']"):
            try:
                import json
                data = json.loads(script.string or "{}")
                candidates = data if isinstance(data, list) else [data]
                for obj in candidates:
                    if isinstance(obj, dict):
                        ho = obj.get("hiringOrganization")
                        if isinstance(ho, dict) and ho.get("name"):
                            company = ho.get("name").strip()
                            break
                        author = obj.get("author")
                        if isinstance(author, dict) and author.get("name"):
                            company = author.get("name").strip()
                            break
                if company:
                    break
            except Exception:
                continue

    # 5) fallback: derive from URL path
    if not company:
        try:
            path = urlparse(job_url).path.strip("/")
            parts = [p for p in path.split("/") if p]
            if parts:
                candidate = parts[1] if len(parts) > 1 else parts[0]
                if candidate and not candidate.isnumeric() and candidate.lower() not in ("punet", "job", "jobs", "companies"):
                    company = candidate.replace("-", " ").replace("_", " ").title()
        except Exception:
            pass

    if not company:
        company = "N/A"

    # ---------- LOCATION ----------
    location = ""
    # badges often contain city
    badges = soup.select("div.inline-flex span, div.inline-flex, span")
    for b in badges:
        text = (b.get_text(" ", strip=True) or "")
        if text and re.search(r"(Prisht|Pristina|Prishtinë|Tirana|Prizren|Gjakov|Ferizaj|Pejë|Mitrov)", text, re.I):
            location = text.strip()
            break

    if not location:
        meta_loc = soup.select_one("meta[name='location'], meta[property='og:locality']")
        if meta_loc and meta_loc.get("content"):
            location = meta_loc.get("content").strip()

    if not location:
        # try breadcrumbs or labeled location areas
        el = soup.find(class_=re.compile(r"(location|job-location|post-location)", re.I))
        if el:
            location = txt(el)

    if not location:
        location = "Kosovo"

    # ---------- DESCRIPTION ----------
    desc = ""
    desc_selectors = [
        ".post-content, .entry-content, .job-description, .content, .single-content, main, article"
    ]
    for sel in desc_selectors:
        el = soup.select_one(sel)
        if el:
            desc = txt(el)
            # remove repeated title/company at start if present
            if desc.startswith(title):
                desc = desc[len(title):].strip()
            break
    if not desc:
        desc = soup.get_text(" ", strip=True)[:4000]

    # ---------- DATES & CONTACTS ----------
    posted_on = ""
    time_el = soup.find("time")
    if time_el and time_el.get("datetime"):
        posted_on = time_el.get("datetime").split("T")[0]
    elif time_el:
        posted_on = time_el.get_text(strip=True)
    else:
        meta_date = soup.select_one("meta[property='article:published_time'], meta[name='date']")
        if meta_date and meta_date.get("content"):
            posted_on = meta_date.get("content").split("T")[0]

    posted_on = posted_on or datetime.utcnow().strftime("%Y-%m-%d")

    page_text = soup.get_text(" ", strip=True)

    email_m = re.search(r'[\w\.-]+@[\w\.-]+\.\w{2,}', page_text)
    phone_m = re.search(r'(\+?\d[\d\s\-\(\)]{5,}\d)', page_text)
    salary_m = re.search(r'(\d{1,3}(?:[.,]\d{3})?\s*(EUR|€|eur))', page_text, re.I)
    jtype_m = re.search(r'(Full[-\s]*time|Part[-\s]*time|Internship|Contract|Temporary|Freelance)', page_text, re.I)
    job_id_m = re.search(r'(\d{4,})', job_url)  # lightweight id attempt


        # ---- CATEGORY (safe extraction) ----
    try:
        cat_el = soup.select_one(".inline-flex")
        if not cat_el:
            cat_el = soup.select_one(".breadcrumb")
        category = txt(cat_el) if cat_el else "General"
    except Exception:
        category = "General"

    return {
        "company": company,
        "position": title,
        "phone_number": phone_m.group(0).strip() if phone_m else "",
        "email": email_m.group(0).strip() if email_m else "",
        "category": category,
        "posting_date": posted_on,
        "scraping_date": datetime.utcnow().strftime("%Y-%m-%d"),
        "location": location,
        "job_link": job_url,
        "description": desc,
        "source": "Telegrafi",
        "website": base_url,
        "salary": salary_m.group(0) if salary_m else "",
        "job_type": jtype_m.group(0).title() if jtype_m else "",
        "jobpost_ad_id": job_id_m.group(0) if job_id_m else ""
    }


# -----------------------------------------
# Main scraper (uses parse_telegrafi_detail)
# -----------------------------------------
def scrape_telegrafi(
    base_url: str = "https://jobs.telegrafi.com",
    max_pages: int = None,
    max_scrolls: int = 20,
    delay: float = 2.0,
    headless: bool = False
):
    """
    Scrape Telegrafi job list + details using Playwright (scrolling + stealth).
    """
    if max_pages is not None:
        max_scrolls = max_pages

    scraped_jobs = []
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless, args=["--no-sandbox", "--disable-setuid-sandbox"])
        context = browser.new_context()
        page = context.new_page()
        stealth_sync(page)

        print("\n=== Loading Telegrafi job list ===")
        try:
            page.goto(base_url, timeout=60000, wait_until="domcontentloaded")
        except Exception as e:
            print("Failed to open base URL:", e)
            browser.close()
            return scraped_jobs

        # wait for probable listing selector
        try:
            page.wait_for_selector("div.relative.group, article, .post, .job, .listing, .jobs-list", timeout=20000)
            print("✓ Listing rendered")
        except PWTimeout:
            print("⚠️ Listing selector not found within timeout. Continuing anyway...")

        # SCROLL TO LOAD MORE JOBS
        last_height = None
        for i in range(max_scrolls):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(delay + random.random() * 0.6)
            new_height = page.evaluate("document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # Parse loaded html and find the real cards
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")

        cards = soup.select("div.relative.group")
        print("Total cards found:", len(cards))

        links = []
        for card in cards:
            a = card.select_one("a[href*='/jobs/']")
            if a and a.get("href"):
                href = a["href"].strip()
                full = urljoin(base_url, href)
                parsed = urlparse(full)
                if parsed.netloc and parsed.netloc.endswith(urlparse(base_url).netloc):
                    if full not in links:
                        links.append(full)

        print("Total job links extracted:", len(links))

        # Load each detail page in same page context
        for job_url in links:
            try:
                time.sleep(0.5 + random.random() * 0.8)
                job = parse_telegrafi_detail(job_url, page, base_url)
                if job:
                    scraped_jobs.append(job)
                    print("✓", job["position"])
                else:
                    print("✗ Failed to parse:", job_url)
            except Exception as e:
                print("Error while parsing", job_url, e)
                continue

        browser.close()
    return scraped_jobs






import time, re, random
from datetime import datetime, UTC
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from playwright_stealth import stealth_sync


# --------------------------
#   DETAIL PAGE PARSER
# --------------------------
from datetime import datetime, timezone

def parse_portalpune_detail(job_url, page, base_url):
    """Parse a PortalPune job detail page using Playwright page."""

    try:
        page.goto(job_url, timeout=60000, wait_until="domcontentloaded")
        time.sleep(1.2)

        soup = BeautifulSoup(page.content(), "html.parser")

        def txt(x):
            return x.get_text(strip=True) if x else ""

        # ===== TITLE =====
        title = txt(soup.select_one(".job-meta h2"))

        # ===== COMPANY =====
        company_el = soup.select_one(".job-overview .card:nth-of-type(1) strong a")
        if not company_el:
            company_el = soup.select_one(".job-overview .card:nth-of-type(1) strong")
        company = txt(company_el)

        # ----------- CATEGORY & LOCATION FIXED HERE -----------
        category = ""
        location = ""

        cards = soup.select(".col-md-12.col-lg-4 .card")
        for c in cards:
            label = c.select_one("span")
            value = c.select_one("strong")

            label_text = label.get_text(strip=True).lower() if label else ""
            value_text = value.get_text(strip=True) if value else ""

            if "categor" in label_text:
                category = value_text

            if "workplace" in label_text or "location" in label_text:
                location = value_text

        # ===== POSTING DATE =====
        posting_el = soup.select(".job-meta ul li strong")
        posted_on = txt(posting_el[0]) if posting_el else ""

        # ===== DESCRIPTION =====
        desc_block = soup.select_one(".description")
        description = desc_block.get_text("\n", strip=True) if desc_block else ""

        # ===== PHONE =====
        phone_m = re.search(r"(?:\+?\d[\d\s\-\/]{6,})", description)

        # ===== EMAIL =====
        email_m = re.search(
            r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
            description
        )

        # ===== SALARY =====
        salary_m = re.search(
            r"(?:\d{2,3}[.,]?\d{0,3}\s?(?:€|eur|euro|Euro|EURO))",
            description,
            re.IGNORECASE
        )

        # ===== JOB TYPE =====
        jtype_m = re.search(
            r"\b(full\s*time|part\s*time|kontratë|orari i plotë|orari i pjesshëm)\b",
            description,
            re.IGNORECASE
        )

        # ===== JOBPOST AD ID =====
        job_id_m = re.search(r"/jobs/(\d+)", job_url)

        # ----- FINAL RETURN -----
        return {
            "company": company or "",
            "position": title or "",
            "phone_number": phone_m.group(0).strip() if phone_m else "",
            "email": email_m.group(0).strip() if email_m else "",
            "category": category,
            "posting_date": posted_on or "",
            "scraping_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "location": location or "",
            "job_link": job_url,
            "description": description,
            "source": "PortalPune",
            "website": base_url,
            "salary": salary_m.group(0) if salary_m else "",
            "job_type": jtype_m.group(0).title() if jtype_m else "",
            "jobpost_ad_id": job_id_m.group(1) if job_id_m else ""
        }

    except Exception as e:
        print(f"Error parsing PortalPune detail {job_url}: {e}")
        return None

# ---------------------------------------------------
#   MAIN SCRAPER – EXACT STYLE OF scrape_telegrafi()
# ---------------------------------------------------
def scrape_portalpune(
    base_url: str = "https://portalpune.com/jobs",
    max_pages: int = None,
    max_scrolls: int = 20,
    delay: float = 2.0,
    headless: bool = False,
):
    """Scrape PortalPune job list + details using Playwright (scroll + stealth)."""

    if max_pages is not None:
        max_scrolls = max_pages

    scraped_jobs = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--no-sandbox", "--disable-setuid-sandbox"]
        )
        context = browser.new_context()
        page = context.new_page()
        stealth_sync(page)

        print("\n=== Loading PortalPune job list ===")

        try:
            page.goto(base_url, timeout=60000, wait_until="domcontentloaded")
        except Exception as e:
            print("❌ Failed to load base:", e)
            browser.close()
            return scraped_jobs

        # Wait for job list
        try:
            page.wait_for_selector(".job-item", timeout=25000)
            print("✓ Listing rendered")
        except PWTimeout:
            print("⚠️ Job list did not appear. Continuing...")

        # --- Scroll to load all jobs ---
        last_height = None
        for i in range(max_scrolls):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(delay + random.random())
            new_height = page.evaluate("document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # Parse HTML
        soup = BeautifulSoup(page.content(), "html.parser")

        # Job cards from your HTML structure
        cards = soup.select(".job-item, .job-item-premium")
        print("Total cards found:", len(cards))

        # Extract job links
        links = []
        for card in cards:
            a = card.select_one("a[href]")
            if a:
                href = a["href"]
                full = urljoin(base_url, href)
                if full not in links:
                    links.append(full)

        print("Total job links extracted:", len(links))

        # --- Parse each job detail page ---
        for job_url in links:
            try:
                time.sleep(0.6 + random.random())
                job = parse_portalpune_detail(job_url, page, base_url)
                if job:
                    scraped_jobs.append(job)
                    print("✓", job["position"])
                else:
                    print("✗ Failed to parse:", job_url)
            except Exception as e:
                print("❌ Error while parsing", job_url, e)

        browser.close()

    return scraped_jobs








def parse_generic_job(job_url, session_req, base_url, source_name):
    """Generic parser for job detail pages"""
    resp = session_req.get(job_url, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, 'html.parser')

    def txt(sel):
        return sel.get_text(strip=True, separator=' ') if sel else ''

    title = txt(soup.select_one('h1, h2.job-title, .job-title, .entry-title, [class*="title"]'))
    company = txt(soup.select_one('.company, .employer, .company-name, [class*="company"]'))
    category = txt(soup.select_one('.category, .breadcrumb, [class*="category"]')) or 'General'
    location = txt(soup.select_one('.location, .job-location, [class*="location"]')) or 'Kosovo'
    posted_on = txt(soup.find('time')) or datetime.utcnow().strftime('%Y-%m-%d')

    desc = txt(soup.select_one('.job-description, .description, .content, .entry-content, [class*="description"]'))
    if not desc:
        desc = soup.get_text(' ', strip=True)[:2000]

    page_text = soup.get_text(' ', strip=True)
    email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w{2,}', page_text)
    phone_match = re.search(r'(\+?\d{2,4}[\s-]?)?(\(?\d{2,4}\)?[\s-]?\d{2,3}[\s-]?\d{2,4})', page_text)
    salary_match = re.search(r'(\d{1,3}(?:[.,]\d{3})?\s*(?:EUR|€|\$|USD))', page_text, re.I)
    job_type_match = re.search(r'(Full[-\s]*time|Part[-\s]*time|Intern|Contract|Freelance)', page_text, re.I)

    jobpost_id = re.search(r'(\d{4,})', job_url)
    colleague_id = None
    profile_link = soup.find('a', href=re.compile(r'/user/|/profile/|/company/|/punedhenes/'))
    if profile_link:
        m = re.search(r'(\d{3,})', profile_link['href'])
        colleague_id = m.group(0) if m else profile_link['href']

    job_data = {
        "company": company or '',
        "position": title or '',
        "phone_number": phone_match.group(0) if phone_match else '',
        "email": email_match.group(0) if email_match else '',
        "category": category,
        "posting_date": posted_on,
        "scraping_date": datetime.utcnow().strftime('%Y-%m-%d'),
        "location": location,
        "job_link": job_url,
        "description": desc,
        "source": source_name,
        "website": base_url,
        "salary": salary_match.group(0) if salary_match else '',
        "job_type": job_type_match.group(0).title() if job_type_match else '',
        "jobpost_ad_id": jobpost_id.group(0) if jobpost_id else '',
        "colleague_user_id": colleague_id or ''
    }

    return job_data


def generate_sample_jobs(source="KosovaJob", website="https://kosovajob.com"):
    """Fallback if scraping fails"""
    return [{
        "company": "Sample Corp",
        "position": "Software Engineer",
        "phone_number": "+383 44 123 456",
        "email": "info@sample.com",
        "category": "IT",
        "posting_date": "2025-11-10",
        "scraping_date": datetime.utcnow().strftime('%Y-%m-%d'),
        "location": "Prishtinë",
        "job_link": f"{website}/sample",
        "description": "Sample job description for testing purposes.",
        "source": source,
        "website": website,
        "salary": "1000 EUR",
        "job_type": "Full-time",
        "jobpost_ad_id": "0001",
        "colleague_user_id": "U123"
    }]



@app.route('/api/jobs')
def get_jobs():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    return jsonify({'jobs': sample_jobs, 'count': len(sample_jobs)})


@app.route('/api/jobs', methods=['DELETE'])
def delete_all_jobs():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    global sample_jobs
    sample_jobs = []
    return jsonify({'success': True, 'message': 'All jobs deleted'})


@app.route('/api/download_excel')
def download_excel():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    # Define headers matching all scraped fields
    headers = [
        "Company", "Position", "Phone Number", "Email", "Category",
        "Posted On", "Data Scraped On", "Location", "Post Web Link",
        "Job Description", "Source", "Website", "Salary",
        "Job Type", "JobPost Ad ID", "Colleague User ID"
    ]

    # Add headers to the worksheet
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092",
                                end_color="366092",
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add job data
    for row, job in enumerate(sample_jobs, 2):
        job_data = [
            job.get('company', ''),
            job.get('position', ''),
            job.get('phone_number', ''),
            job.get('email', ''),
            job.get('category', ''),
            job.get('posting_date', ''),
            job.get('scraping_date', ''),
            job.get('location', ''),
            job.get('job_link', ''),
            job.get('description', ''),
            job.get('source', ''),
            job.get('website', ''),
            job.get('salary', ''),
            job.get('job_type', ''),
            job.get('jobpost_ad_id', ''),
            job.get('colleague_user_id', '')
        ]

        for col, value in enumerate(job_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="left",
                                       vertical="top",
                                       wrap_text=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO object
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Generate filename with timestamp
    filename = f"job_listings_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Serve React App in production
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react_app(path):
    # Check if the request is for a static file
    if path and os.path.exists(os.path.join('client/dist', path)):
        return send_from_directory('client/dist', path)
    # Otherwise, serve index.html for client-side routing
    if os.path.exists('client/dist/index.html'):
        return send_from_directory('client/dist', 'index.html')
    # Fallback for development
    return jsonify({'message': 'Frontend not built. Run: cd client && npm run build'}), 404


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)
